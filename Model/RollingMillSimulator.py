from math import *
from RollingMill import RollingMill
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class RollingMillSimulator(RollingMill):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.time_log = [0]  # Лог отображения нынешнего иммитируемого времени
        self.temperature_log = [self.StartTemp]  # Лог изменения температуры сляба
        self.length_log = [self.L]  # Лог изменения длины сляба
        self.height_log = [self.h_0]  # Лог толщины сляба(перед началом прокатки)(мм)
        self.LeftCap = [0]  # Левый концевик
        self.RightCap = [0]  # Правый концевик
        self.x_log = [0]  # Лог начальной координаты сляба
        self.x1_log = [self.L]  # Лог конечной координаты сляба
        self.pyrometr_1 = [self.TempV]  # Лог пирометра перед валками
        self.pyrometr_2 = [self.TempV]  # Лог пирометра после валков
        self.gap_log = [self.CurrentS]  # Лог раствора валков(мм)
        self.speed_V = [0]  # Лог скорости варщения валков(об/c)
        self.speed_V0 = [0]  # Лог скорости вращения рольгангов до валков(об/c)
        self.speed_V1 = [0]  # Лог скорости вращения рольгангов после валков(об/c)
        self.effort_log = [0]  # Лог усилия прокаткатки(кН)
        self.moment_log = [0]  # Лог момента прокаткатки(кН*м)
        self.power_log = [0]  # Лог мощности прокатки(кВт)
        self.Gap_feedbackLog = [0]  # Лог флага обратной свзяи о выхождении раствора на заданную уставку 
        self.Speed_V_feedbackLog = [0]  # Лог флага обратной свзяи о выхождении скорости валков на заданную уставку
        self.time_step = 0.1  # Шаг времени

    def roughness(self, number, Range) -> float:
        'Генерация случайного отклонения на +- n процентов от заданного числа для симуляции неровностей сляба'
        five_percent = number * Range
        random_deviation = random.uniform(-five_percent, five_percent)
        return random_deviation
        
    def linear_interpolation(self, start, end, steps) -> float:
        "Линейная интерполяция, возвращающая шаг смещения величины"
        if steps <= 0:
            raise ValueError("Количество шагов должно быть положительным")
        step_size = (end - start) / steps
        return step_size
    
    def clear_logs(self):
        self.time_log = [0]  # Лог отображения нынешнего иммитируемого времени
        self.temperature_log = [self.StartTemp]  # Лог изменения температуры сляба
        self.length_log = [self.L]  # Лог изменения длины сляба
        self.height_log = [self.h_0]  # Лог толщины сляба(перед началом прокатки)(мм)
        self.LeftCap = [0]  # Левый концевик
        self.RightCap = [0]  # Правый концевик
        self.x_log = [0]  # Лог начальной координаты сляба
        self.x1_log = [self.L]  # Лог конечной координаты сляба
        self.pyrometr_1 = [self.TempV]  # Лог пирометра перед валками
        self.pyrometr_2 = [self.TempV]  # Лог пирометра после валков
        self.gap_log = [self.CurrentS]  # Лог раствора валков(мм)
        self.speed_V = [0]  # Лог скорости варщения валков(об/c)
        self.speed_V0 = [0]  # Лог скорости вращения рольгангов до валков(об/c)
        self.speed_V1 = [0]  # Лог скорости вращения рольгангов после валков(об/c)
        self.effort_log = [0]  # Лог усилия прокаткатки(кН)
        self.moment_log = [0]  # Лог момента прокаткатки(кН*м)
        self.power_log = [0]  # Лог мощности прокатки(кВт)
        self.Gap_feedbackLog = [0]  # Лог флага обратной свзяи о выхождении раствора на заданную уставку 
        self.Speed_V_feedbackLog = [0]  # Лог флага обратной свзяи о выхождении скорости валков на заданную уставку

    
    def save_logs_to_excel(self, filename="rolling_log.xlsx"):
        "Сохраняет логи в XLSX файл с русскими названиями столбцов и форматированием"
        wb = Workbook()
        ws = wb.active
        ws.title = "Логи прокатки"
        
        headers = [
            'Время (с)',
            'Пирометр 1 (°C)',
            'Пирометр 2 (°C)', 
            'Температура сляба (°C)',
            'Усилие прокатки (кН)',
            'Зазор валков (мм)',
            'Скорость валков (мм/c)',
            'Скорость рольгангов 0 (мм/c)',
            'Скорость рольгангов 1 (мм/c)',
            'Левый концевик',
            'Правый концевик',
            'Момент прокатки (кН*м)',
            'Мощность прокатки (кВт)',
            'Флаг зазора',
            'Флаг скорости',
            'Координата X (мм)',
            'Координата X1 (мм)',
            'Длина сляба (мм)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx in range(len(self.time_log)):
            row_data = [
                self.time_log[row_idx],
                self.pyrometr_1[row_idx],
                self.pyrometr_2[row_idx],
                self.temperature_log[row_idx],
                self.effort_log[row_idx],
                self.gap_log[row_idx],
                self.speed_V[row_idx],
                self.speed_V0[row_idx],
                self.speed_V1[row_idx],
                self.LeftCap[row_idx],
                self.RightCap[row_idx],
                self.moment_log[row_idx],
                self.power_log[row_idx],
                self.Gap_feedbackLog[row_idx],
                self.Speed_V_feedbackLog[row_idx],
                self.x_log[row_idx],
                self.x1_log[row_idx],
                self.length_log[row_idx]
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
        
        column_widths = [
            12,  # Время (с)
            18,  # Пирометр 1 (°C)
            18,  # Пирометр 2 (°C)
            24,  # Температура сляба (°C)
            20,  # Усилие прокатки (кН)
            18,  # Зазор валков (мм)
            24,  # Скорость валков (мм/c)
            28,  # Скорость рольгангов 0 (мм/c)
            28,  # Скорость рольгангов 1 (мм/c)
            18,  # Левый концевик
            18,  # Правый концевик
            24,  # Момент прокатки (кН*м)
            22,  # Мощность прокатки (кВт)
            15,  # Флаг зазора
            16,  # Флаг скорости
            20,  # Координата X (мм)
            20,  # Координата X1 (мм)
            20   # Длина сляба (мм)
        ]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(filename)
        print(f"Логи успешно сохранены в файл: {filename}")
    
    def _update_logs(self, time, gap, speed_V, temp, pyrometr_1, pyrometr_2, pos_x, pos_x1, speed_V0, speed_V1, length, effort, moment, power, LeftCap, RightCap, Gap_feedback, Speed_V_feedback):
        "Обновление внутренних логов без записи в файл"
        self.time_log.append(time)
        self.gap_log.append(gap)
        self.speed_V.append(speed_V)
        self.pyrometr_1.append(pyrometr_1)
        self.pyrometr_2.append(pyrometr_2)
        self.temperature_log.append(temp)
        self.x_log.append(pos_x)
        self.x1_log.append(pos_x1)
        self.speed_V0.append(speed_V0)
        self.speed_V1.append(speed_V1)
        self.length_log.append(length)
        self.effort_log.append(effort)
        self.moment_log.append(moment)
        self.power_log.append(power)
        self.LeftCap.append(LeftCap)
        self.RightCap.append(RightCap)
        self.Gap_feedbackLog.append(Gap_feedback)
        self.Speed_V_feedbackLog.append(Speed_V_feedback)

    def _Gap_Valk_(self, Roll_pos, Dir_of_rot_valk): 
        Gap_flag = self.Gap_feedbackLog[-1]
        current_time = self.time_log[-1]
        CurrentS = self.gap_log[-1]
        self.h_0 = self.h_0 if CurrentS == 350 else CurrentS
        speed_V0 = self.speed_V0[-1] 
        speed_V1 = self.speed_V1[-1] 
        self.Dir_of_rot = Dir_of_rot_valk
        current_temp = self.temperature_log[-1]
        self.h_1 = Roll_pos
        
        gap_change_per_ms = self.VS * self.time_step
        self.S = Roll_pos
        target_gap = self.S
        time_gap = (abs(self.S - CurrentS)) / (self.VS)
        final_drop = self.TempDrBPass(T0 = self.temperature_log[-1],Time = time_gap,width =self.b,height=self.h_0)
        final_temp = current_temp - final_drop
        temp_drop_per_ms = ((current_temp - final_temp) / time_gap) * self.time_step

        while CurrentS != target_gap:
            CurrentS = min(CurrentS + gap_change_per_ms, target_gap) if CurrentS < target_gap else max(CurrentS - gap_change_per_ms, target_gap)
            if CurrentS == self.S:
                Gap_flag = 1
            else:
                Gap_flag = 0
            current_temp = max(current_temp - temp_drop_per_ms, final_temp)
            Pyro1 = self.TempV + self.roughness(self.TempV, 0.07)
            Pyro2 = self.TempV + self.roughness(self.TempV, 0.07)
            current_time += self.time_step
            self._update_logs(time=round(current_time, 1), 
                              gap=round(CurrentS, 2), 
                              speed_V=0, 
                              temp=round(current_temp, 2),
                              pyrometr_1=round(Pyro1, 2),
                              pyrometr_2=round(Pyro2, 2), 
                              pos_x=round(self.x_log[-1], 2),
                              pos_x1=round(self.x1_log[-1]), 
                              speed_V0=round(speed_V0, 2), 
                              speed_V1=round(speed_V1, 2), 
                              length=round(self.length_log[-1], 2),
                              effort=0,
                              moment=0,
                              power=0,
                              LeftCap=self.LeftCap[-1],
                              RightCap=self.RightCap[-1],
                              Gap_feedback=Gap_flag,
                              Speed_V_feedback=self.Speed_V_feedbackLog[-1]) 
        return self._get_current_state()

    def _Accel_Valk_(self,Num_of_revol_rolls,Dir_of_rot_L_rolg,Dir_of_rot_R_rolg):
        current_time = self.time_log[-1]
        current_speed = self.speed_V[-1]
        current_temp = self.temperature_log[-1]
        Speed_V_flag = self.Speed_V_feedbackLog[-1]
        
        # self.V_Valk_Per = (2 * pi * self.DV/2 * Num_of_revol_rolls) / 60
        time_accel = ((Num_of_revol_rolls) / (self.accel))
        final_drop = self.TempDrBPass(T0 = self.temperature_log[-1],Time = time_accel,width =self.b,height=self.h_0)
        final_temp = current_temp - final_drop
        temp_drop_per_ms = ((current_temp - final_temp) / time_accel) * self.time_step
        
        while current_speed != Num_of_revol_rolls:
            current_speed = min(current_speed + self.accel * self.time_step,Num_of_revol_rolls) 
            current_time += self.time_step
            current_temp = max(current_temp - temp_drop_per_ms, final_temp)
            Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
            Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
            if current_speed == Num_of_revol_rolls:
                Speed_V_flag = 1
            self._update_logs(time=round(current_time,1), 
                              gap=round(self.gap_log[-1],2), 
                              speed_V=round(current_speed,2), 
                              temp=round(current_temp,2), 
                              pyrometr_1=round(Pyro1,2), 
                              pyrometr_2=round(Pyro2,2),
                              pos_x=round(self.x_log[-1],2),
                              pos_x1=round(self.x1_log[-1],2), 
                              speed_V0=0, 
                              speed_V1=0, 
                              length=round(self.length_log[-1],2),
                              effort=0,
                              moment=0,
                              power=0,
                              LeftCap=self.LeftCap[-1],
                              RightCap=self.RightCap[-1],
                              Gap_feedback = self.Gap_feedbackLog[-1],
                              Speed_V_feedback = Speed_V_flag)
        return {
            'Time': self.time_log,
            'Pyro1': self.pyrometr_1,
            'Pyro2': self.pyrometr_2,
            'Power': self.power_log,
            'Gap': self.gap_log,
            'VRPM': self.speed_V,
            'V0RPM': self.speed_V0,
            'V1RPM': self.speed_V1,
            'Moment': self.moment_log,
            'Pressure': self.effort_log,
            'StartCap': self.LeftCap,
            'EndCap':self.RightCap,
            'Gap_feedback': self.Gap_feedbackLog,
            'Speed_feedback': self.Speed_V_feedbackLog,
            'Length' : self.length_log
        }


    def _Approching_to_Roll_(self,Dir_of_rot,Num_of_revol_0rollg,Num_of_revol_1rollg):
        "Проход сляба к валкам"
        current_pos_x = self.x_log[-1]
        current_pos_x1 = self.x1_log[-1]
        current_time = self.time_log[-1]
        current_temp = self.temperature_log[-1]
        Left_Cap = self.LeftCap[-1]
        Right_Cap = self.RightCap[-1]
        speed_V0 = 0
        speed_V1 = 0

        self.Dir_of_rot = Dir_of_rot
        # self.V0 = (2 * pi * self.DR/2 * Num_of_revol_0rollg) / 60
        # self.V1 = (2 * pi * self.DR/2 * Num_of_revol_1rollg) / 60
        Offset = sqrt((self.DV/2)**2 - (self.DV/2 - ((self.h_0 - self.h_1)/2))**2)

        if self.Dir_of_rot == 0:
            time_accel = ((self.V0) / (self.accel))
            S1 = ((self.accel) * time_accel**2)/2
            S2 = (self.d1 + self.d/2 - self.L - Offset) - S1 
            time_max_speed = (S2 / (Num_of_revol_0rollg))
            time_move = (time_accel + time_max_speed)
        else:
            time_accel = ((self.V1) / (self.accel))
            S1 = ((self.accel) * time_accel**2)/2
            S2 = (self.d1 + Offset) - S1              
            time_max_speed = (S2 / (Num_of_revol_0rollg))
            time_move = (time_accel + time_max_speed) 

        final_drop = self.TempDrBPass(T0 = self.temperature_log[-1],Time = time_move,width =self.b,height=self.h_0)

        final_temp = current_temp - final_drop
        temp_drop_per_ms = ((current_temp - final_temp) / time_move) * self.time_step

        if self.Dir_of_rot == 0:
            while current_pos_x1 != self.d1 + self.d/2 - Offset:
                current_temp = max(current_temp - temp_drop_per_ms, final_temp)
                speed_V0 = min(speed_V0 + self.accel * self.time_step, Num_of_revol_0rollg)
                speed_V1 = min(speed_V1 + self.accel * self.time_step, Num_of_revol_1rollg)
                current_pos_x = min(current_pos_x + speed_V0 * self.time_step, self.d1 + self.d/2 - Offset - self.length_log[-1]) 
                current_pos_x1 = min(current_pos_x1 + speed_V0 * self.time_step, self.d1 + self.d/2 - Offset)
                if current_pos_x1 >= self.LeftStopCap and current_pos_x <= self.LeftStopCap:
                    Left_Cap = 1
                else:
                    Left_Cap = 0
                if current_pos_x1 >= 2000 and current_pos_x <= 2000:
                    Pyro1 = current_temp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
                Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                                gap=round(self.gap_log[-1],2), 
                                speed_V=round(self.speed_V[-1],2), 
                                temp=round(current_temp,2), 
                                pyrometr_1=round(Pyro1,2),
                                pyrometr_2=round(Pyro2,2), 
                                pos_x=round(current_pos_x,2),
                                pos_x1=round(current_pos_x1,2), 
                                speed_V0=round(speed_V0,2), 
                                speed_V1=round(speed_V1,2), 
                                length=round(self.length_log[-1],2),
                                effort=0,
                                moment=0,
                                power=0,
                                LeftCap=Left_Cap,
                                RightCap=self.RightCap[-1],
                                Gap_feedback = self.Gap_feedbackLog[-1],
                                Speed_V_feedback = self.Speed_V_feedbackLog[-1])
        else:
            while current_pos_x != self.d1 + self.d/2 + Offset:
                current_temp = max(current_temp - temp_drop_per_ms, final_temp)
                speed_V0 = min(speed_V0 + self.accel * self.time_step, Num_of_revol_0rollg)
                speed_V1 = min(speed_V1 + self.accel * self.time_step, Num_of_revol_1rollg)
                current_pos_x = max(current_pos_x - speed_V1 * self.time_step, self.d1 + self.d/2 + Offset)
                current_pos_x1 = max(current_pos_x1 - speed_V1 * self.time_step, self.d1+self.d/2 + Offset + self.length_log[-1])
                if current_pos_x <= self.RightStopCap and current_pos_x1 >= self.RightStopCap:
                    Right_Cap = 1
                else:
                    Right_Cap = 0
                if current_pos_x <= 2700 and current_pos_x1 >= 2700:
                    Pyro2 = current_temp
                else:
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                                gap=round(self.gap_log[-1],2), 
                                speed_V=round(self.speed_V[-1],2), 
                                temp=round(current_temp,2), 
                                pyrometr_1=round(Pyro1,2),
                                pyrometr_2=round(Pyro2,2), 
                                pos_x=round(current_pos_x,2),
                                pos_x1=round(current_pos_x1,2), 
                                speed_V0=round(speed_V0,2), 
                                speed_V1=round(speed_V1,2), 
                                length=round(self.length_log[-1],2),
                                effort=0,
                                moment=0,
                                power=0,
                                LeftCap=self.LeftCap[-1],
                                RightCap=Right_Cap,
                                Gap_feedback = self.Gap_feedbackLog[-1],
                                Speed_V_feedback = self.Speed_V_feedbackLog[-1])
        return {
            'Time': self.time_log,
            'Pyro1': self.pyrometr_1,
            'Pyro2': self.pyrometr_2,
            'Power': self.power_log,
            'Gap': self.gap_log,
            'VRPM': self.speed_V,
            'V0RPM': self.speed_V0,
            'V1RPM': self.speed_V1,
            'Moment': self.moment_log,
            'Pressure': self.effort_log,
            'StartCap': self.LeftCap,
            'EndCap':self.RightCap,
            'Gap_feedback': self.Gap_feedbackLog,
            'Speed_feedback': self.Speed_V_feedbackLog,
            'Length' : self.length_log
        }


    def _simulate_rolling_pass(self):
        "Симуляция прохода сляба через валки"
        current_pos_x = self.x_log[-1]
        current_pos_x1 = self.x1_log[-1]
        current_length = self.length_log[-1]
        current_time = self.time_log[-1]

        h_0 = self.h_0
        h_1 = self.S
        RelDef = self.RelDef(h_0,h_1)
        Length_coef = self.h_0 / self.h_1
 
        RelDef = self.RelDef(h_0,h_1)
        ContactArcLen = self.ContactArcLen(self.DV,h_0=h_0,h_1=h_1)
        DefResistance = self.DefResistance(RelDef=RelDef,LK=ContactArcLen,V=self.speed_V[-1],CurrentTemp=self.temperature_log[-1],SteelGrade=self.SteelGrade)
        AvrgPressure = self.AvrgPressure(DefResistance=DefResistance,LK=ContactArcLen,h_0=h_0,h_1=h_1)
        Effort = self.Effort(LK=ContactArcLen,b=self.b,AvrgPressure=AvrgPressure)
        Moment = self.Moment(LK=ContactArcLen,h_0=h_0,h_1=h_1,Effort=Effort/1000)
        Power = self.Power(Moment,self.speed_V[-1],self.DV)
        SpeedOfRolling = self.SpeedOfRolling(DV=self.DV,V=self.speed_V[-1])
        TempDrDConRoll = self.TempDrDConRoll(DV=self.DV,h_0=h_0,h_1=h_1,Temp=self.temperature_log[-1],SpeedOfRolling=SpeedOfRolling)
        TempDrPlDeform = self.TempDrPlDeform(DefResistance=DefResistance,h_0=h_0,h_1=h_1)
        GenTemp = self.GenTemp(Temp=self.temperature_log[-1],TempDrDConRoll=TempDrDConRoll,TempDrPlDeform=TempDrPlDeform,TempDrBPass=0) 
          
        if self.Dir_of_rot == 0:
            while current_pos_x1 != self.d1 + self.d/2:
                current_pos_x1 = min(current_pos_x1 + self.speed_V[-1]  * self.time_step,self.d1 + self.d/2)
                current_pos_x = min(current_pos_x + self.speed_V[-1] * self.time_step,self.d1 + self.d/2 -current_length) 
                current_length = current_pos_x1 - current_pos_x               
                Effort += self.roughness(Effort,0.03)
                Moment += self.roughness(Moment,0.03)
                Power += self.roughness(Power,0.03)
                if current_pos_x1 >= 2700 and current_pos_x <= 2700:
                    Pyro2 = GenTemp
                else:
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                
                if current_pos_x <= 2000:
                    Pyro1 = GenTemp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
                
                if current_pos_x1 >= self.RightStopCap and current_pos_x <= self.RightStopCap:
                    RightCap = 1
                else:
                    RightCap = 0
                
                if current_pos_x1 >= self.LeftStopCap and current_pos_x <= self.LeftStopCap:
                    LeftCap = 1
                else:
                    LeftCap = 0 
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                        gap=round(self.gap_log[-1],2), 
                        speed_V=round(self.speed_V[-1],2), 
                        temp=round(GenTemp,2), 
                        pyrometr_1=round(Pyro1,2),
                        pyrometr_2=round(Pyro2,2), 
                        pos_x=round(current_pos_x,2), 
                        pos_x1=round(current_pos_x1,2), 
                        speed_V0=round(self.speed_V0[-1],2), 
                        speed_V1=round(self.speed_V1[-1],2), 
                        length=round(current_length,2),
                        effort=round(Effort/1000,2),
                        moment=round(Moment/1000,2),
                        power= round(Power/1000,2),
                        LeftCap=LeftCap,
                        RightCap=RightCap,
                        Gap_feedback = self.Gap_feedbackLog[-1],
                        Speed_V_feedback = self.Speed_V_feedbackLog[-1]
                        ) 
            
            while current_pos_x <= self.d1 + self.d/2:
                current_pos_x1 = current_pos_x1 + self.speed_V[-1] * Length_coef * self.time_step
                current_pos_x = current_pos_x + self.speed_V[-1] * self.time_step
                current_length = current_pos_x1 - current_pos_x
                Effort += self.roughness(Effort,0.03)
                Moment += self.roughness(Moment,0.03)
                Power += self.roughness(Power,0.03)
                
                if current_pos_x1 >= 2700 and current_pos_x <= 2700:
                    Pyro2 = GenTemp
                else:
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                
                if current_pos_x <= 2000:
                    Pyro1 = GenTemp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
                
                if current_pos_x1 >= self.RightStopCap and current_pos_x <= self.RightStopCap:
                    RightCap = 1
                else:
                    RightCap = 0
                
                if current_pos_x1 >= self.LeftStopCap and current_pos_x <= self.LeftStopCap:
                    LeftCap = 1
                else:
                    LeftCap = 0 
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                        gap=round(self.gap_log[-1],2), 
                        speed_V=round(self.speed_V[-1],2), 
                        temp=round(GenTemp,2), 
                        pyrometr_1=round(Pyro1,2),
                        pyrometr_2=round(Pyro2,2), 
                        pos_x=round(current_pos_x,2), 
                        pos_x1=round(current_pos_x1,2), 
                        speed_V0=round(self.speed_V0[-1],2), 
                        speed_V1=round(self.speed_V1[-1],2), 
                        length=round(current_length,2),
                        effort=round(Effort/1000,2),
                        moment=round(Moment/1000,2),
                        power= round(Power/1000,2),
                        LeftCap=LeftCap,
                        RightCap=RightCap,
                        Gap_feedback = self.Gap_feedbackLog[-1],
                        Speed_V_feedback = self.Speed_V_feedbackLog[-1]
                        ) 
        else:
            while current_pos_x != self.d1 + self.d/2:
                current_pos_x = max(current_pos_x - self.speed_V[-1]  * self.time_step,self.d1 + self.d/2)
                current_pos_x1 = max(current_pos_x1 - self.speed_V[-1] * self.time_step,self.d1 + self.d/2 + current_length)
                current_length = current_pos_x1 - current_pos_x
                Effort += self.roughness(Effort,0.03)
                Moment += self.roughness(Moment,0.03)
                Power += self.roughness(Power,0.03)
                if current_pos_x1 <= 2700 and current_pos_x >= 2700:
                    Pyro2 = GenTemp
                else:  
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                if current_pos_x <= 2000:
                    Pyro1 = GenTemp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07) 

                if current_pos_x <= self.RightStopCap and current_pos_x1 >= self.RightStopCap:
                    RightCap = 1
                else:
                    RightCap = 0 

                if current_pos_x <= self.LeftStopCap and current_pos_x1 >= self.LeftStopCap:
                    LeftCap = 1
                else:
                    LeftCap = 0  
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                        gap=round(self.gap_log[-1],2), 
                        speed_V=round(self.speed_V[-1],2), 
                        temp=round(GenTemp,2), 
                        pyrometr_1=round(Pyro1,2),
                        pyrometr_2=round(Pyro2,2), 
                        pos_x=round(current_pos_x,2), 
                        pos_x1=round(current_pos_x1,2), 
                        speed_V0=round(self.speed_V0[-1],2), 
                        speed_V1=round(self.speed_V1[-1],2), 
                        length=round(current_length,2),
                        effort=round(Effort/1000,2),
                        moment=round(Moment/1000,2),
                        power= round(Power/1000,2),
                        LeftCap=LeftCap,
                        RightCap=RightCap,
                        Gap_feedback = self.Gap_feedbackLog[-1],
                        Speed_V_feedback = self.Speed_V_feedbackLog[-1]
                        )        
            
            while current_pos_x1 >= self.d1 + self.d/2:
                current_pos_x = current_pos_x - self.speed_V[-1] * Length_coef * self.time_step
                current_pos_x1 = current_pos_x1 - self.speed_V[-1] * self.time_step
                current_length = current_pos_x1 - current_pos_x
                Effort += self.roughness(Effort,0.03)
                Moment += self.roughness(Moment,0.03)
                Power += self.roughness(Power,0.03)    
                if current_pos_x1 <= 2700 and current_pos_x >= 2700:
                    Pyro2 = GenTemp
                else:  
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                if current_pos_x <= 2000:
                    Pyro1 = GenTemp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07) 

                if current_pos_x <= self.RightStopCap and current_pos_x1 >= self.RightStopCap:
                    RightCap = 1
                else:
                    RightCap = 0 

                if current_pos_x <= self.LeftStopCap and current_pos_x1 >= self.LeftStopCap:
                    LeftCap = 1
                else:
                    LeftCap = 0          
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                        gap=round(self.gap_log[-1],2), 
                        speed_V=round(self.speed_V[-1],2), 
                        temp=round(GenTemp,2), 
                        pyrometr_1=round(Pyro1,2),
                        pyrometr_2=round(Pyro2,2), 
                        pos_x=round(current_pos_x,2), 
                        pos_x1=round(current_pos_x1,2), 
                        speed_V0=round(self.speed_V0[-1],2), 
                        speed_V1=round(self.speed_V1[-1],2), 
                        length=round(current_length,2),
                        effort=round(Effort/1000,2),
                        moment=round(Moment/1000,2),
                        power= round(Power/1000,2),
                        LeftCap=LeftCap,
                        RightCap=RightCap,
                        Gap_feedback = self.Gap_feedbackLog[-1],
                        Speed_V_feedback = self.Speed_V_feedbackLog[-1]
                        ) 
        return {
            'Time': self.time_log,
            'Pyro1': self.pyrometr_1,
            'Pyro2': self.pyrometr_2,
            'Power': self.power_log,
            'Gap': self.gap_log,
            'VRPM': self.speed_V,
            'V0RPM': self.speed_V0,
            'V1RPM': self.speed_V1,
            'Moment': self.moment_log,
            'Pressure': self.effort_log,
            'StartCap': self.LeftCap,
            'EndCap':self.RightCap,
            'Gap_feedback': self.Gap_feedbackLog,
            'Speed_feedback': self.Speed_V_feedbackLog,
            'Length' : self.length_log
        }

    def _simulate_exit_from_rolls(self):
        "Симуляция дохода сляба до концевика"
        Speed_V_flag = self.Speed_V_feedbackLog[-1]
        current_time = self.time_log[-1]
        current_temp = self.temperature_log[-1]
        LeftCap = self.LeftCap[-1]
        RightCap = self.RightCap[-1]
        x = self.x_log[-1]
        x1 = self.x1_log[-1]
        #1.Рассчет падения температуры  
        distance_to_cover = (self.d/2 + self.d2) - self.x_log[-1]
        time_first_cycle = distance_to_cover / self.speed_V1[-1]
        time_brake_speed = self.speed_V[-1] / self.accel
        time_brake_V0 = self.speed_V0[-1] / self.accel  
        time_brake_V1 = self.speed_V1[-1] / self.accel
        time_second_cycle = max(time_brake_speed, time_brake_V0, time_brake_V1)
        total_time = time_first_cycle + time_second_cycle
        final_drop = self.TempDrBPass(T0 = self.temperature_log[-1],Time = total_time,width =self.b,height=self.h_0)
        final_temp = current_temp - final_drop
        temp_drop_per_ms = ((current_temp - final_temp) / total_time) * self.time_step
        #2.Доход сляба до конечного концевика
        if self.Dir_of_rot == 0:
            while self.x1_log[-1] != self.RightStopCap :
                current_temp -= temp_drop_per_ms
                Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
                x = min(self.x_log[-1] + self.speed_V1[-1] * self.time_step,self.RightStopCap - self.length_log[-1])
                x1 = min(self.x1_log[-1] + self.speed_V1[-1] * self.time_step,self.RightStopCap)
                
                if x1 >= 2700 and x <= 2700:
                    Pyro2 = current_temp
                else:
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                
                if x1 >= self.RightStopCap and x <= self.RightStopCap:
                    RightCap = 1
                else:
                    RightCap = 0
                
                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                                gap=round(self.gap_log[-1],2), 
                                speed_V=round(self.speed_V[-1],2), 
                                temp=round(current_temp,2), 
                                pyrometr_1=round(Pyro1,2),
                                pyrometr_2=round(Pyro2,2), 
                                pos_x=round(x,2), 
                                pos_x1=round(x1,2), 
                                speed_V0=round(self.speed_V0[-1],2), 
                                speed_V1=round(self.speed_V1[-1],2), 
                                length=round(self.length_log[-1],2),
                                effort=0,
                                moment=0,
                                power=0,
                                LeftCap=LeftCap,
                                RightCap=RightCap,
                                Gap_feedback = self.Gap_feedbackLog[-1],
                                Speed_V_feedback = self.Speed_V_feedbackLog[-1])
        else:
            while x != self.LeftStopCap:
                current_temp -= temp_drop_per_ms
                Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                x1 = max(self.x1_log[-1] - self.speed_V0[-1] * self.time_step,self.LeftStopCap + self.length_log[-1])
                x = max(self.x_log[-1] - self.speed_V0[-1] * self.time_step,self.LeftStopCap)

                if x <= 2000 and x1 >= 2000:
                    Pyro1 = current_temp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07)

                if x < self.LeftStopCap and x1 > self.LeftStopCap:
                    LeftCap = 1
                else:
                    LeftCap = 0 

                current_time += self.time_step
                self._update_logs(time=round(current_time,1), 
                    gap=round(self.gap_log[-1],2), 
                    speed_V=round(self.speed_V[-1],2), 
                    temp=round(current_temp,2), 
                    pyrometr_1=round(Pyro1,2),
                    pyrometr_2=round(Pyro2,2), 
                    pos_x=round(x,2), 
                    pos_x1=round(x1,2), 
                    speed_V0=round(self.speed_V0[-1],2), 
                    speed_V1=round(self.speed_V1[-1],2), 
                    length=round(self.length_log[-1],2),
                    effort=0,
                    moment=0,
                    power=0,
                    LeftCap=LeftCap,
                    RightCap=RightCap,
                    Gap_feedback = self.Gap_feedbackLog[-1],
                    Speed_V_feedback = self.Speed_V_feedbackLog[-1])
                
        #3.Замедление рольгангов и валков до 0 скорости
        current_speed = self.speed_V[-1]
        current_V0 = self.speed_V0[-1]
        current_V1 = self.speed_V1[-1]
        current_temp = self.temperature_log[-1]
        if self.Dir_of_rot == 0:
            while current_speed > 0 or current_V0 > 0 or current_V1 > 0:
                current_speed = max(current_speed - self.accel * self.time_step,0)
                current_V0 = max(current_V0 - self.accel * self.time_step,0)
                current_V1 = max(current_V1 - self.accel * self.time_step,0)
                x = self.x_log[-1] + self.speed_V1[-1] * self.time_step
                x1 = self.x1_log[-1] + self.speed_V1[-1] * self.time_step
                
                if current_speed != self.V_Valk_Per:
                    Speed_V_flag = 0
                if current_speed == 0:
                    Speed_V_flag = 1
                current_temp = max(current_temp - temp_drop_per_ms, final_temp)
                
                if x1 >= self.RightStopCap and x <= self.RightStopCap:
                    RightCap = 1
                else:
                    RightCap = 0
                
                if self.x_log[-1] <= 2700 and self.x1_log[-1] >= 2700:
                    Pyro2 = current_temp
                else:
                    Pyro2 = self.TempV + self.roughness(self.TempV,0.07)
                current_time += self.time_step
                self._update_logs(time=round(current_time,1),
                                gap=round(self.gap_log[-1],2),
                                speed_V=round(current_speed,2),
                                temp=round(current_temp,2),
                                pyrometr_1=round(Pyro1,2),
                                pyrometr_2=round(Pyro2,2),
                                pos_x=round(x,2),
                                pos_x1= round(x1,2),
                                speed_V0=round(current_V0,2),
                                speed_V1=round(current_V1,2),
                                length=round(self.length_log[-1],2),
                                effort=0,
                                moment=0,
                                power=0,
                                LeftCap=self.LeftCap[-1],
                                RightCap=RightCap,
                                Gap_feedback = self.Gap_feedbackLog[-1],
                                Speed_V_feedback = Speed_V_flag)
        else:
            while current_speed > 0 or current_V0 > 0 or current_V1 > 0:
                current_speed = max(current_speed - self.accel * self.time_step,0)
                current_V0 = max(current_V0 - self.accel * self.time_step,0)
                current_V1 = max(current_V1 - self.accel * self.time_step,0)
                x = self.x_log[-1] - self.speed_V1[-1] * self.time_step
                x1 = self.x1_log[-1] - self.speed_V1[-1] * self.time_step
                                
                if x <= self.LeftStopCap and x1 >= self.LeftStopCap:
                    LeftCap = 1
                else:
                    LeftCap = 0  
                
                if current_speed != self.V_Valk_Per:
                    Speed_V_flag = 0
                
                if current_speed == 0:
                    Speed_V_flag = 1
                
                current_temp = max(current_temp - temp_drop_per_ms, final_temp)
                
                if self.x_log[-1] <= 2000 and self.x1_log[-1] >= 2000:
                    Pyro1 = current_temp
                else:
                    Pyro1 = self.TempV + self.roughness(self.TempV,0.07)
                current_time += self.time_step
                self._update_logs(time=round(current_time,1),
                                gap=round(self.gap_log[-1],2),
                                speed_V=round(current_speed,2),
                                temp=round(current_temp,2),
                                pyrometr_1=round(Pyro1,2),
                                pyrometr_2=round(Pyro2,2),
                                pos_x=round(x,2),
                                pos_x1= round(x1,2),
                                speed_V0=round(current_V0,2),
                                speed_V1=round(current_V1,2),
                                length=round(self.length_log[-1],2),
                                effort=0,
                                moment=0,
                                power=0,
                                LeftCap=LeftCap,
                                RightCap=self.RightCap[-1],
                                Gap_feedback = self.Gap_feedbackLog[-1],
                                Speed_V_feedback = Speed_V_flag)
        return {
            'Time': self.time_log,
            'Pyro1': self.pyrometr_1,
            'Pyro2': self.pyrometr_2,
            'Power': self.power_log,
            'Gap': self.gap_log,
            'VRPM': self.speed_V,
            'V0RPM': self.speed_V0,
            'V1RPM': self.speed_V1,
            'Moment': self.moment_log,
            'Pressure': self.effort_log,
            'StartCap': self.LeftCap,
            'EndCap':self.RightCap,
            'Gap_feedback': self.Gap_feedbackLog,
            'Speed_feedback': self.Speed_V_feedbackLog,
            'Length' : self.length_log
        }

    def Alarm_stop(self):
        "Аварийная остановка прокатного стана"
        current_time = self.time_log[-1]
        current_gap = self.gap_log[-1]
        while self.speed_V[-1] > 0 or self.speed_V0[-1] > 0 or self.speed_V1[-1] > 0 or self.gap_log[-1] != 350:
            current_speed = max(self.speed_V[-1] - self.accel * self.time_step,0)
            current_V0 = max(self.speed_V0[-1] - self.accel * self.time_step,0)
            current_V1 = max(self.speed_V1[-1] - self.accel * self.time_step,0)
            current_gap =  min(current_gap + self.VS, 350) if current_gap < 350 else max(current_gap - self.VS, 350)
            if current_gap == 350:
                GapCap = 1
            else:
                GapCap = 0
            current_time += self.time_step
            self._update_logs(time=round(current_time,1),
                              gap=round(current_gap,2),
                              speed_V=round(current_speed,2),
                              temp=round(self.temperature_log[-1]),
                              pyrometr_1=round(self.TempV,2),
                              pyrometr_2=round(self.TempV,2),
                              pos_x=round(self.x_log[-1],2),
                              pos_x1= round(self.x1_log[-1],2),
                              speed_V0=round(current_V0,2),
                              speed_V1=round(current_V1,2),
                              length=round(self.length_log[-1],2),
                              effort=0,
                              moment=0,
                              power=0,
                              LeftCap=self.LeftCap[-1],
                              RightCap=self.RightCap[-1],
                              Gap_feedback = GapCap,
                              Speed_V_feedback = 0)
        return {
            'Time': self.time_log,
            'Pyro1': self.pyrometr_1,
            'Pyro2': self.pyrometr_2,
            'Power': self.power_log,
            'Gap': self.gap_log,
            'VRPM': self.speed_V,
            'V0RPM': self.speed_V0,
            'V1RPM': self.speed_V1,
            'Moment': self.moment_log,
            'Pressure': self.effort_log,
            'StartCap': self.LeftCap,
            'EndCap':self.RightCap,
            'Gap_feedback':self.Gap_feedbackLog,
            'Speed_feedback':self.Speed_V_feedbackLog,
            'Length':self.length_log
        } 
    
    def _get_current_state(self):
        "Возвращает текущее состояние всех логов"
        return {
            'Time': self.time_log,
            'Pyro1': self.pyrometr_1,
            'Pyro2': self.pyrometr_2,
            'Power': self.power_log,
            'Gap': self.gap_log,
            'VRPM': self.speed_V,
            'V0RPM': self.speed_V0,
            'V1RPM': self.speed_V1,
            'Moment': self.moment_log,
            'Pressure': self.effort_log,
            'StartCap': self.LeftCap,
            'EndCap': self.RightCap,
            'Gap_feedback': self.Gap_feedbackLog,
            'Speed_feedback': self.Speed_V_feedbackLog,
            'Length': self.length_log
        }

    def Init(self, Length_slab, Width_slab, Thikness_slab, Temperature_slab, Material_slab, Diametr_roll, Material_roll):
        self.CurrentS = 350
        self.TempV = 28
        self.L = Length_slab
        self.b = Width_slab
        self.h_0 = Thikness_slab
        self.StartTemp = Temperature_slab
        self.DV = Diametr_roll
        self.MV = Material_roll
        self.SteelGrade = Material_slab
        self.time_log = [0]
        self.temperature_log = [self.StartTemp]
        self.length_log = [self.L]
        self.height_log = [self.h_0]
        self.LeftCap = [0]
        self.RightCap = [0]
        self.x_log = [0]
        self.x1_log = [self.L]
        self.pyrometr_1 = [self.TempV]
        self.pyrometr_2 = [self.TempV]
        self.gap_log = [self.CurrentS]
        self.speed_V = [0]
        self.speed_V0 = [0]
        self.speed_V1 = [0]
        self.effort_log = [0]
        self.moment_log = [0]
        self.power_log = [0]
        self.Gap_feedbackLog = [0]
        self.Speed_V_feedbackLog = [0]
        self.time_step = 0.1
        self.DV = Diametr_roll
        self.R = self.DV/2
        self.DR = 40
        self.d1 = 2130.0
        self.d2 = 2130.0
        self.d = 440.0
        self.MS = 'Austenitic steel'
        self.VS = 100.0
        self.LeftStopCap = 850 
        self.RightStopCap = 3850  

if __name__ == "__main__":
    simulator = RollingMillSimulator(
        L=0, b=0, h_0=0, S=0, StartTemp=0,
        DV=0, MV=0, MS=0, OutTemp=0, DR=0, SteelGrade=0,
        V0=0, V1=0, VS=0, Dir_of_rot=0,
        d1=0, d2=0, d=0, V_Valk_Per=0, StartS=0,
    )
    simulator.Init(Length_slab=300, Width_slab=250, Thikness_slab=350, Temperature_slab=1200, Material_slab='Ст3сп', Diametr_roll=300, Material_roll='Сталь')
    

    simulator._Gap_Valk_(330, 0)
    simulator._Accel_Valk_(200, 0, 0)
    simulator._Approching_to_Roll_(0, 200, 212)
    simulator._simulate_rolling_pass()
    simulator._simulate_exit_from_rolls()

    simulator._Gap_Valk_(300, 1)
    simulator._Accel_Valk_(200, 1, 1)
    simulator._Approching_to_Roll_(1, 220, 200)
    simulator._simulate_rolling_pass()
    simulator._simulate_exit_from_rolls()
    
    simulator.save_logs_to_excel("my_logs.xlsx")