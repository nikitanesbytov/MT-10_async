import asyncio
import struct
from datetime import datetime

import asyncpg
from pymodbus.server import StartAsyncTcpServer
from pymodbus.datastore import ModbusSequentialDataBlock
from pymodbus.datastore import ModbusSlaveContext, ModbusServerContext

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from RollingMillSimulator import RollingMillSimulator


def float_to_regs(value: float):
    """Преобразует float в два WORD регистра (big-endian)."""
    b = struct.pack('>f', float(value))
    return [int.from_bytes(b[:2], 'big'), int.from_bytes(b[2:], 'big')]


def regs_to_float(reg1: int, reg2: int) -> float:
    """Преобразует два WORD регистра обратно в float (big-endian)."""
    b1 = (reg1 >> 8) & 0xFF
    b2 = reg1 & 0xFF
    b3 = (reg2 >> 8) & 0xFF
    b4 = reg2 & 0xFF
    return struct.unpack('>f', bytes([b1, b2, b3, b4]))[0]

class AsyncModbusServer:
    def __init__(self):
        # Настройка области регистров Holding Registers
        total_registers = 33
        initial_values = [0] * total_registers
        # Адрес 1, длина total_registers
        self.hr_data_combined = ModbusSequentialDataBlock(1, initial_values)
        store = ModbusSlaveContext(hr=self.hr_data_combined)
        self.context = ModbusServerContext(slaves=store, single=True)

        self.stop_monitoring = False
        self.simulator: RollingMillSimulator | None = None
        self.initialized = False

        # Состояние "этапа" прокатки
        self.counter = 0
        self.counter2 = 0
        self.nex_idx = 0
        self.prev_total_steps = 0
        self.status_code = 0 # 1-ожидание инициализации,2-ожидание переключателя старт,3-ожидание команды старта, 4-проход выполняется

        # Синхронизация асинхронных задач
        self.simulation_lock = asyncio.Lock()
        self.simulation_in_progress = False

        # --- Excel логирование ---
        self.excel_wb: Workbook | None = None
        self.excel_ws = None
        self.excel_filename: str | None = None
        self.excel_next_row: int | None = None

    # ===================== Excel-помощники =====================

    def _create_new_excel_workbook(self, last_row):
        """
        Создаёт новый Excel-файл для текущего процесса прокатки.
        last_row — запись из таблицы slabs (asyncpg.Record) с начальными параметрами.
        """
        # Имя файла — текущая дата/время окончания инициализации
        now = datetime.now()
        self.excel_filename = now.strftime("%H:%M:%S_%d.%m.%Y.xlsx")

        self.excel_wb = Workbook()
        self.excel_ws = self.excel_wb.active
        ws = self.excel_ws
        ws.title = "Лог прокатки"

        # --- Блок начальных параметров ---
        ws.cell(row=1, column=1, value="Параметр").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Значение").font = Font(bold=True)

        params = [
            ("length_slab",      "Длина сляба, мм"),
            ("width_slab",       "Ширина сляба, мм"),
            ("thikness_slab",    "Толщина сляба, мм"),
            ("temperature_slab", "Температура сляба, °C"),
            ("material_slab",    "Марка стали"),
            ("diametr_roll",     "Диаметр валков, мм"),
            ("material_roll",    "Материал валков"),
        ]

        row = 2
        for key, title in params:
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=2, value=last_row[key])
            row += 1

        # Пустая строка
        row += 1

        # --- Заголовки для построчных данных (то, что идёт в регистры) ---
        headers = [
            "Время (с)",
            "Пирометр 1 (°C)",
            "Пирометр 2 (°C)",
            "Давление (кН)",
            "Раствор (мм)",
            "Скорость валков (об/с)",
            "Скорость левой группы рольгангов (об/с)",
            "Скорость правой группы рольгангов (об/с)",
            "Момент прокатки (кН·м)",
            "Мощность прокатки (кВт)",
        ]

        header_row = row
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Первая строка для данных
        self.excel_next_row = header_row + 1

        # Можно чуть растянуть ширину колонок
        for col in range(1, len(headers) + 1):
            col_letter = ws.cell(row=header_row, column=col).column_letter
            ws.column_dimensions[col_letter].width = 14

        # Сохранить структуру файла
        self.excel_wb.save(self.excel_filename)

    def _log_step_to_excel(self, sim_data: dict, idx: int):
        """
        Записать один шаг симуляции (индекс idx) в Excel.
        Логируются только те значения, которые реально идут в регистры.
        """
        if not (self.excel_wb and self.excel_ws and self.excel_next_row):
            return  # логирование ещё не настроено

        ws = self.excel_ws
        r = self.excel_next_row

        def get(key, default=0.0):
            if key not in sim_data:
                return default
            val = sim_data[key]
            if isinstance(val, list):
                if 0 <= idx < len(val):
                    return val[idx]
                return default
            return val

        row_data = [
            get("Time", 0.0),
            get("Pyro1", 0.0),
            get("Pyro2", 0.0),
            get("Pressure", 0.0),
            get("Gap", 0.0),
            get("VRPM", 0.0),
            get("V0RPM", 0.0),
            get("V1RPM", 0.0),
            get("Moment", 0.0),
            get("Power", 0.0),
        ]

        for c, value in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=value)

        self.excel_next_row += 1

        # Для простоты сохраняем файл каждый шаг (можно оптимизировать при необходимости)
        self.excel_wb.save(self.excel_filename)

    # ===================== Логика аварийной остановки =====================

    async def alarm_stop(self, diff: int):
        """Асинхронное выполнение последовательности аварийной остановки."""
        async with self.simulation_lock:
            self.simulation_in_progress = True
            try:
                alarm_data = self.simulator.Alarm_stop()
                await self._write_alarm_data_to_registers(alarm_data, diff)
            finally:
                self.simulation_in_progress = False

    async def _write_alarm_data_to_registers(self, alarm_data: dict, diff: int):
        """Асинхронно записывает данные аварийной остановки в регистры и Excel."""
        total_steps = len(alarm_data['Time'])
        self.nex_idx += diff

        while self.nex_idx != total_steps and not self.stop_monitoring:
            self._write_single_step_to_registers_sync(alarm_data, self.nex_idx)
            self.nex_idx += 1
            await asyncio.sleep(0.1)

    # ===================== Основная инициализация =====================

    async def start_init_from_registers(self):
        """Асинхронная инициализация из БД."""
        async with self.simulation_lock:
            self.simulation_in_progress = True
            self.status_code = 1
            self.hr_data_combined.setValues(33, [self.status_code])
            
            try:
                reg32 = self.hr_data_combined.getValues(32, 1)[0]
                new_reg32 = reg32

                # Сброс выходных регистров и установка начального зазора 350
                self.hr_data_combined.setValues(12, [0] * 21)
                gap_regs = float_to_regs(350)
                self.hr_data_combined.setValues(18, gap_regs)

                while not self.stop_monitoring:
                    try:
                        conn = await asyncio.wait_for(
                            asyncpg.connect(
                                host="localhost",
                                database="postgres",
                                user="postgres",
                                password="postgres",
                                port="5432"
                            ),
                            timeout=1.0
                        )
                        try:
                            # Проверяем количество записей и удаляем старые
                            count_result = await conn.fetchrow(
                                "SELECT COUNT(*) AS count FROM slabs"
                            )
                            if count_result and count_result['count'] > 10:
                                # Оставляем только 3 последние записи
                                await conn.execute(
                                    """
                                    DELETE FROM slabs 
                                    WHERE id NOT IN (
                                        SELECT id FROM slabs 
                                        ORDER BY id DESC 
                                        LIMIT 3
                                    )
                                    """
                                )

                            last_row = await conn.fetchrow(
                                "SELECT * FROM slabs ORDER BY id DESC LIMIT 1"
                            )
                            if last_row and not last_row['is_used']:
                                sim = RollingMillSimulator(
                                    L=0, b=0, h_0=0, S=0, StartTemp=0,
                                    DV=0, MV=0, MS=0, OutTemp=0, DR=0, SteelGrade=0,
                                    V0=0, V1=0, VS=0, Dir_of_rot=0,
                                    d1=0, d2=0, d=0, V_Valk_Per=0, StartS=350
                                )

                                ms_clean = (last_row['material_slab'] or "").replace(' ', '')
                                sim.Init(
                                    Length_slab=last_row['length_slab'],
                                    Width_slab=last_row['width_slab'],
                                    Thikness_slab=last_row['thikness_slab'],
                                    Temperature_slab=last_row['temperature_slab'],
                                    Material_slab=ms_clean,
                                    Diametr_roll=last_row['diametr_roll'],
                                    Material_roll=last_row['material_roll']
                                )

                                self.simulator = sim
                                self.initialized = True
                                new_reg32 = reg32 | 0x10
                                self.hr_data_combined.setValues(32, [new_reg32])

                                await conn.execute(
                                    "UPDATE public.slabs SET is_used = TRUE WHERE id = $1",
                                    last_row['id']
                                )

                                # Создаём новый Excel-файл для этого запуска
                                self._create_new_excel_workbook(last_row)

                                self.nex_idx = 0
                                self.prev_total_steps = 0
                                self.counter = 0
                                self.counter2 = 0
                                break
                        finally:
                            await conn.close()
                    except asyncio.TimeoutError:
                        # Просто пробуем ещё раз
                        continue
            finally:
                self.simulation_in_progress = False

    # ===================== Запуск Modbus-сервера =====================

    async def run_server(self, IP: str, port: int):
        """Асинхронный запуск Modbus сервера."""
        try:
            await StartAsyncTcpServer(context=self.context, address=(IP, port))
        finally:
            self.stop_monitoring = True

    # ===================== Запись данных симуляции =====================

    async def write_simulation_data_to_registers(self, sim_data: dict):
        """Запись данных симуляции в регистры и Excel."""
        async with self.simulation_lock:
            self.simulation_in_progress = True
            try:
                total_steps = len(sim_data['Time'])
                diff = total_steps - self.prev_total_steps
                self.prev_total_steps = total_steps

                while self.nex_idx != total_steps and not self.stop_monitoring:
                    self._write_single_step_to_registers_sync(sim_data, self.nex_idx)
                    self.nex_idx += 1
                    diff -= 1

                    # Читаем управляющий регистр (адрес 9, индекс 8 относительно начала 1)
                    regs = self.hr_data_combined.getValues(1, 11)
                    reg8 = regs[8]
                    Alarm = bool(reg8 & 0x08)
                    Reset_alarm = bool(reg8 & 0x01)

                    if Alarm:
                        await self.alarm_stop(diff)
                        self.initialized = False
                        return

                    if Reset_alarm:
                        await self.start_init_from_registers()
                        self.initialized = False
                        return

                    await asyncio.sleep(0.1)
            finally:
                self.simulation_in_progress = False

    def _write_single_step_to_registers_sync(self, sim_data: dict, idx: int):
        """
        Синхронно записывает данные одного шага симуляции в регистры
        и добавляет такую же строку в Excel.
        """
        keys = [
            'Pyro1', 'Pyro2', 'Pressure', 'Gap', 'VRPM', 'V0RPM', 'V1RPM',
            'Moment', 'Power', 'Time'
        ]
        regs = []
        for k in keys:
            v = sim_data[k][idx] if isinstance(sim_data[k], list) else sim_data[k]
            regs.extend(float_to_regs(v))

       
        self.hr_data_combined.setValues(12, regs)

        # Формирование битовых флагов
        flags = 0
        StartCap_val = sim_data['StartCap'][idx] if isinstance(sim_data['StartCap'], list) else sim_data['StartCap']
        EndCap_val = sim_data['EndCap'][idx] if isinstance(sim_data['EndCap'], list) else sim_data['EndCap']
        Gap_feedback_val = sim_data['Gap_feedback'][idx] if isinstance(sim_data['Gap_feedback'], list) else sim_data['Gap_feedback']
        Speed_feedback_val = sim_data['Speed_feedback'][idx] if isinstance(sim_data['Speed_feedback'], list) else sim_data['Speed_feedback']

        if StartCap_val:
            flags |= 0x01
        if EndCap_val:
            flags |= 0x02
        if Gap_feedback_val:
            flags |= 0x04
        if Speed_feedback_val:
            flags |= 0x08

        self.hr_data_combined.setValues(32, [flags])

        # Логирование шага в Excel
        self._log_step_to_excel(sim_data, idx)

    # ===================== Мониторинг управляющих регистров =====================

    async def monitor_registers(self):
        """Мониторинг регистров с проверкой блокировки."""
        while not self.stop_monitoring:
            if self.simulation_in_progress or self.simulation_lock.locked():
                await asyncio.sleep(0.1)
                continue

            if not self.initialized:
                await asyncio.sleep(0.1)
                continue
            
            regs = self.hr_data_combined.getValues(1, 9)
            reg8 = regs[8]

            Reset_alarm = bool(reg8 & 0x01)
            Dir_of_rot_rolg = bool(reg8 & 0x02)
            Dir_of_rot_valk = Dir_of_rot_rolg
            Dir_of_rot = Dir_of_rot_rolg
            Hand_Mode = bool(reg8 & 0x04)
            Start = bool(reg8 & 0x10)
            Start_Gap = bool(reg8 & 0x20)
            Start_Accel = bool(reg8 & 0x40)
            Start_Roll = bool(reg8 & 0x80)
            Start_Switch = bool(reg8 & 0x100)

            if Reset_alarm:
                self.initialized = False
                await self.start_init_from_registers()
                continue
            if Start_Switch:
                if Start and not self.simulation_in_progress:
                    # 1. Установка зазора
                    self.status_code = 4
                    self.hr_data_combined.setValues(33, [self.status_code])
                    if Start_Gap and self.counter == 0 and self.counter2 < 2:
                        Roll_pos = regs_to_float(regs[2], regs[3])
                        sim_result = self.simulator._Gap_Valk_(Roll_pos, Dir_of_rot_valk)
                        await self.write_simulation_data_to_registers(sim_result)
                        self.counter = 1
                        self.counter2 += 1

                    # 2. Разгон валков
                    if Start_Accel and self.counter == 1 and self.counter2 < 2:
                        Num_of_revol_rolls = regs_to_float(regs[0], regs[1])
                        sim_result = self.simulator._Accel_Valk_(Num_of_revol_rolls, Dir_of_rot_rolg, Dir_of_rot_rolg)
                        await self.write_simulation_data_to_registers(sim_result)
                        self.counter = 2
                        self.counter2 += 1

                    # 3. Подход, проход и выход из валков
                    if Start_Roll and self.counter == 2 and self.counter2 <= 2:
                        Num_of_revol_0rollg = regs_to_float(regs[4], regs[5])
                        Num_of_revol_1rollg = regs_to_float(regs[6], regs[7])

                        sim_result = self.simulator._Approching_to_Roll_(
                            Dir_of_rot,
                            Num_of_revol_0rollg,
                            Num_of_revol_1rollg,
                        )
                        await self.write_simulation_data_to_registers(sim_result)

                        sim_result = self.simulator._simulate_rolling_pass()
                        await self.write_simulation_data_to_registers(sim_result)

                        sim_result = self.simulator._simulate_exit_from_rolls()
                        await self.write_simulation_data_to_registers(sim_result)

                        self.counter2 += 1
                else:
                    self.status_code = 3
                    self.hr_data_combined.setValues(33, [self.status_code])
                    self.counter = 0
                    self.counter2 = 0
            else:
                self.counter = 0
                self.counter2 = 0
                self.nex_idx = 0
                self.prev_total_steps = 0
                self.status_code = 2
                self.hr_data_combined.setValues(33, [self.status_code])
                RollingMillSimulator.clear_logs(self.simulator)
            await asyncio.sleep(0.1)

# ===================== Точка входа =====================

async def main():
    server = AsyncModbusServer()

    server_task = asyncio.create_task(server.run_server("192.168.0.99", 55000))

    # Дадим серверу подняться
    await asyncio.sleep(0.1)

    # Первичная инициализация (создаёт Excel-файл для первой заготовки)
    await server.start_init_from_registers()

    await asyncio.gather(
        server.monitor_registers(),
        server_task
    )

if __name__ == "__main__":
    asyncio.run(main())
